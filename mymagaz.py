import openpyxl
import pymysql
import bs4
import requests
from config import host, user, password, db_name

mymagaz_sqp = []
mymagaz_sq = []

# Формирование ТехГранд
book = openpyxl.open('TehGrand.xlsx', read_only=True)
sheet = book.active
for row in range(2, sheet.max_row + 1):
    sku = str(sheet[row][2].value)
    sku = sku.replace(' ', '')
    quantity = sheet[row][3].value
    pr = sheet[row][4].value
    if pr >= 3000:
        price = (pr * 1.1)
    else:
        price = (pr + 300)
    tg = []
    tg.append(quantity)
    tg.append(price)
    tg.append(sku)
    mymagaz_sqp.append(tg)
print(mymagaz_sqp)
print("ТехГранд - ОК")

# Формирование УкрКлимат
book = openpyxl.open('UkrKlimat.xlsx', read_only=True)
sheet = book.active
for row in range(3, sheet.max_row + 1):
    sku = str(sheet[row][8].value)
    sku = sku.replace(' ', '')
    quantity = sheet[row][7].value
    for x, y in ('<10', '10'), ('<5', '5'), ('нет', '0'):
        quantity = quantity.replace(x, y)
    pr = sheet[row][6].value
    if pr < 3000:
        pri = pr + 300
    else:
        pri = pr
    uk = []
    uk.append(quantity)
    uk.append(pri)
    uk.append(sku)
    mymagaz_sqp.append(uk)
print(mymagaz_sqp)
print("УкрКлимат - ОК")


# #Формирование БСХ
# book = openpyxl.open('bsh.xlsx', read_only=True)
# sheet = book.active
# for row in range(3, sheet.max_row+1):
#     sku = sheet[row][0].value
#     quantity = sheet[row][6].value
#     for x, y in ('Больше 5', '5'), ('Есть', '2'), ('Ожидается', '0'), ('Нет', '0'):
#         quantity = quantity.replace(x, y)
#     bsh = []
#     bsh.append(quantity)
#     bsh.append(sku)
#     mymagaz_sq.append(bsh)
# print(mymagaz_sq)
# print("BSH - ОК")
#
# #Формирование Teka
# book = openpyxl.open('teka.xlsx', read_only=True)
# sheet = book.active
# for row in range(20, sheet.max_row+1):
#     sku = sheet[row][0].value
#     quantity = str(sheet[row][2].value)
#     pr = sheet[row][3].value
#     quantity = quantity.replace('> 10', '10')
#     teka = []
#     teka.append(quantity)
#     teka.append(pr)
#     teka.append(sku)
#     mymagaz_sqp.append(teka)
# print(mymagaz_sqp)
# print("TEKA - ОК")
#
# #Запись Franke в CSV
# book = openpyxl.open('franke.xlsx', read_only=True)
# sheet = book.active
# for row in range(2, sheet.max_row+1):
#     sku = sheet[row][1].value
#     quantity = str(sheet[row][3].value)
#     quantity = quantity.replace('наявність по запиту', '0')
#     franke = []
#     franke.append(quantity)
#     franke.append(sku)
#     mymagaz_sq.append(franke)
# print(mymagaz_sq)
# print("Franke - ОК")
#
#
#
#Формирование Karcher
# url = 'http://xml.pricecop.net/ua.kaercher.com/price.xml'
# url_code = requests.get(url)
# soup = bs4.BeautifulSoup(url_code.text, 'lxml')
# quantity = soup.find_all("stock")
# sku = soup.find_all("id")
# karcher = []
# for i in range (0, len(quantity)):
#     row = [quantity[i].get_text(), sku[i].get_text()]
#     karcher.append(row)
#     mymagaz_sq.append(row)
# print(karcher)
# print(mymagaz_sq)
# print("Karcher - ОК")

try:
    connection = pymysql.connect(
        host=host,
        port=3306,
        user=user,
        password=password,
        database=db_name,
        cursorclass=pymysql.cursors.DictCursor
    )
    print("successfully connected...")
    print("#" * 20)

    try:
        with connection.cursor() as cursor:
            update_000tg = "UPDATE `oc_product` SET `quantity`= 0 WHERE `mpn` = 'TG'"
            update_000uk = "UPDATE `oc_product` SET `quantity`= 0 WHERE `mpn` = 'UK'"
            update_000karcher = "UPDATE `oc_product` SET `quantity`= 0 WHERE `manufacturer_id` = '50'"
            update_query_sqp = "UPDATE oc_product SET quantity = %s, price = %s WHERE sku = %s"
            update_query_sq = "UPDATE oc_product SET quantity = %s WHERE sku = %s"
            cursor.execute(update_000tg)
            cursor.execute(update_000uk)
            cursor.execute(update_000karcher)
            print('Quantity = 000"TehGrand"')
            print('Quantity = 000"UkrKlimat"')
            print('Quantity = 000"Karcher"')
            cursor.executemany(update_query_sqp, mymagaz_sqp)
            print('sqp - OK')
            cursor.executemany(update_query_sq, mymagaz_sq)
            print('sq - OK')
            connection.commit()
            print('All Updated')

    except Exception as ex:
        print("Error...")
        print(ex)

    finally:
        connection.close()
        print("Connection close")
except Exception as ex:
    print("Connection refused...")
    print(ex)

finally:
    print("All Complete")
